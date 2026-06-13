#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Analyze Coupang product spreadsheet exports for XHS content."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import zipfile
from pathlib import Path
from statistics import median
from typing import Any
from xml.etree import ElementTree


COLUMN_ALIASES = {
    "id": ["id", "商品id", "商品ID", "product_id", "product id"],
    "sku": ["sku", "SKU", "货号"],
    "image": ["主图", "main_image", "image", "图片", "商品主图"],
    "name": ["商品名称", "名称", "product_name", "name", "标题"],
    "delivery": ["配送类型", "配送方式", "火箭类型", "delivery", "delivery_type", "物流"],
    "price": ["价格", "price", "售价"],
    "reviews": ["评论数", "评价数", "review_count", "reviews"],
    "rating": ["评分", "rating", "score"],
    "clicks": ["点击量", "clicks", "click_count"],
    "sales": ["销量", "sales", "sold", "销售量"],
    "conversion": ["转化率", "conversion", "conversion_rate"],
    "brand": ["品牌", "brand"],
    "category": ["品类", "category"],
}

MISSING_LABEL = "未知"


def normalize_header(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", "", value).lower()


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = (
        text.replace(",", "")
        .replace("%", "")
        .replace("₩", "")
        .replace("￥", "")
        .replace("¥", "")
        .strip()
    )
    try:
        number = float(text)
    except ValueError:
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def xlsx_column_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    index = 0
    for char in letters:
        index = index * 26 + (ord(char.upper()) - ord("A") + 1)
    return index - 1


def read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        data = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ElementTree.fromstring(data)
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    strings: list[str] = []
    for item in root.findall("x:si", namespace):
        parts = [node.text or "" for node in item.findall(".//x:t", namespace)]
        strings.append("".join(parts))
    return strings


def read_xlsx_rows_without_openpyxl(path: Path) -> list[dict[str, Any]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = read_shared_strings(archive)
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        ns = {
            "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
        }
        first_sheet = workbook.find("x:sheets/x:sheet", ns)
        if first_sheet is None:
            return []
        rel_id = first_sheet.attrib.get(f"{{{ns['r']}}}id")
        target = None
        for rel in rels.findall("rel:Relationship", ns):
            if rel.attrib.get("Id") == rel_id:
                target = rel.attrib.get("Target")
                break
        if not target:
            return []
        sheet_path = "xl/" + target.lstrip("/")
        sheet_path = sheet_path.replace("xl/worksheets/../", "xl/")
        sheet = ElementTree.fromstring(archive.read(sheet_path))

    rows: list[list[Any]] = []
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    for row in sheet.findall(".//x:sheetData/x:row", namespace):
        values: list[Any] = []
        for cell in row.findall("x:c", namespace):
            ref = cell.attrib.get("r", "")
            column_index = xlsx_column_index(ref)
            while len(values) <= column_index:
                values.append("")
            cell_type = cell.attrib.get("t")
            value_node = cell.find("x:v", namespace)
            inline_node = cell.find("x:is/x:t", namespace)
            raw_value = value_node.text if value_node is not None else None
            if cell_type == "s" and raw_value is not None:
                value = shared_strings[int(raw_value)]
            elif cell_type == "inlineStr" and inline_node is not None:
                value = inline_node.text or ""
            else:
                value = raw_value or ""
            values[column_index] = value
        rows.append(values)

    if not rows:
        return []
    headers = [normalize_header(value) for value in rows[0]]
    output: list[dict[str, Any]] = []
    for row in rows[1:]:
        output.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return output


def read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return read_xlsx_rows_without_openpyxl(path)

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [normalize_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return rows


def read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix == ".xlsx":
        return read_xlsx_rows(path)
    raise SystemExit(f"Unsupported spreadsheet format: {suffix}. Use .xlsx or .csv.")


def detect_columns(headers: list[str]) -> dict[str, str | None]:
    lookup = {normalize_key(header): header for header in headers}
    detected: dict[str, str | None] = {}
    for field, aliases in COLUMN_ALIASES.items():
        detected[field] = None
        for alias in aliases:
            hit = lookup.get(normalize_key(alias))
            if hit:
                detected[field] = hit
                break
    return detected


def values(rows: list[dict[str, Any]], column: str | None) -> list[float]:
    if not column:
        return []
    return [num for row in rows if (num := to_number(row.get(column))) is not None]


def stats(numbers: list[float]) -> dict[str, float] | None:
    if not numbers:
        return None
    ordered = sorted(numbers)
    return {
        "min": ordered[0],
        "median": median(ordered),
        "max": ordered[-1],
        "avg": round(sum(ordered) / len(ordered), 4),
    }


def top_rows(
    rows: list[dict[str, Any]],
    detected: dict[str, str | None],
    key: str,
    limit: int = 8,
) -> list[dict[str, Any]]:
    column = detected.get(key)
    if not column:
        return []
    scored = [(to_number(row.get(column)) or 0, row) for row in rows]
    scored.sort(key=lambda item: item[0], reverse=True)
    output = []
    for _, row in scored[:limit]:
        output.append({
            "id": row.get(detected.get("id") or "", ""),
            "sku": row.get(detected.get("sku") or "", ""),
            "name": row.get(detected.get("name") or "", ""),
            "delivery": row.get(detected.get("delivery") or "", ""),
            "price": row.get(detected.get("price") or "", ""),
            "reviews": row.get(detected.get("reviews") or "", ""),
            "rating": row.get(detected.get("rating") or "", ""),
            "clicks": row.get(detected.get("clicks") or "", ""),
            "sales": row.get(detected.get("sales") or "", ""),
            "conversion": row.get(detected.get("conversion") or "", ""),
        })
    return output


def count_values(rows: list[dict[str, Any]], column: str | None) -> dict[str, int]:
    if not column:
        return {}
    counts: dict[str, int] = {}
    for row in rows:
        label = str(row.get(column) or MISSING_LABEL).strip() or MISSING_LABEL
        counts[label] = counts.get(label, 0) + 1
    return dict(sorted(counts.items(), key=lambda item: item[1], reverse=True))


def analyze(path: Path) -> dict[str, Any]:
    rows = read_rows(path)
    headers = list(rows[0].keys()) if rows else []
    detected = detect_columns(headers)
    numeric = {
        key: stats(values(rows, detected.get(key)))
        for key in ["price", "reviews", "rating", "clicks", "sales", "conversion"]
    }

    return {
        "file": str(path),
        "row_count": len(rows),
        "detected_columns": detected,
        "missing_fields": [key for key, column in detected.items() if column is None],
        "numeric_stats": numeric,
        "delivery_counts": count_values(rows, detected.get("delivery")),
        "brand_counts": count_values(rows, detected.get("brand")),
        "category_counts": count_values(rows, detected.get("category")),
        "top_by_sales": top_rows(rows, detected, "sales"),
        "top_by_clicks": top_rows(rows, detected, "clicks"),
        "top_by_reviews": top_rows(rows, detected, "reviews"),
        "top_by_rating": top_rows(rows, detected, "rating"),
        "top_by_conversion": top_rows(rows, detected, "conversion"),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_file")
    parser.add_argument("--out", help="Write JSON analysis to this path.")
    args = parser.parse_args()

    result = analyze(Path(args.input_file))
    text = json.dumps(result, ensure_ascii=False, indent=2)
    if args.out:
        Path(args.out).write_text(text, encoding="utf-8")
    else:
        print(text)


if __name__ == "__main__":
    main()


